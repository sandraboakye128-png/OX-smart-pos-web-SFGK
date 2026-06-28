from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, session
from functools import wraps
from services.product_service import get_all_products as get_all_products_service
import os
import uuid
import threading

# ---------- IMPORT PURCHASE SERVICES ----------
from services.purchase_service import (
    add_purchase,
    get_all_purchases,
    get_purchases_by_date_range,
    get_product_suggestions,
    get_category_suggestions,
    update_product
)

# ---------- IMPORT DASHBOARD SERVICES ----------
from services.dashboard_service import (
    get_today_sales,
    get_today_profit,
    get_total_products,
    get_low_stock_products,
    get_top_products,
    get_sales_history
)

# ---------- IMPORT PRODUCT DELETION SERVICES ----------
from services.product_service import (
    delete_product_keep_history,
    delete_product_clean_all,
    delete_batch,
    delete_batch_clean_all
)

# ---------- IMPORT SALES SERVICES ----------
from services.sales_service import (
    get_products_for_sale,
    get_batches_for_product,
    get_batch_by_id,
    create_multi_sale,
    update_product_stock,
    import_sale_bulk
)

# ---------- IMPORT RECEIPT SERVICE ----------
from services.receipt_service import generate_receipt_multi

# ---------- IMPORT AUTH SERVICES ----------
from services.auth_service import (
    login_user,
    create_user,
    admin_exists,
    count_admins,
    get_all_users,
    update_user_role,
    delete_user
)

# ---------- IMPORT ARCHIVE SERVICES ----------
from services.product_service import get_deleted_products, restore_archive

# ---------- DATABASE CONNECTION ----------
from database.db import get_connection
# We'll also need DATABASE_URL – if not exported, read from env
try:
    from database.db import DATABASE_URL
except ImportError:
    DATABASE_URL = os.getenv("DATABASE_URL")

# ---------- PDF GENERATION ----------
import io
from datetime import datetime, date, timedelta
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import cm

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "temporary-dev-key")

# ===================== LICENSE / TRIAL SYSTEM =====================
import json
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LICENSE_FILE = os.path.join(BASE_DIR, "license.json")
TRIAL_DAYS = 30
MASTER_KEY = "OXSMART-1234-KEY"

# ===================== IMPORT JOB TRACKING (DATABASE VERSION) =====================
def init_import_jobs_table():
    """Create the import_jobs table if it does not exist."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS import_jobs (
            job_id UUID PRIMARY KEY,
            status TEXT NOT NULL,
            total INTEGER DEFAULT 0,
            processed INTEGER DEFAULT 0,
            errors JSONB DEFAULT '[]',
            result JSONB,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # In case the table already existed with result as TEXT, alter it to JSONB
    try:
        cursor.execute("ALTER TABLE import_jobs ALTER COLUMN result TYPE JSONB USING result::jsonb")
    except Exception as e:
        # Column might not exist or already JSONB – ignore
        pass
    conn.commit()
    conn.close()
    print("✅ import_jobs table ensured.")

def update_job_progress(job_id, **kwargs):
    """Update job progress in the database."""
    conn = get_connection()
    cursor = conn.cursor()
    set_parts = []
    params = []
    for key, value in kwargs.items():
        if key in ('errors', 'result'):
            set_parts.append(f"{key} = %s::jsonb")
            params.append(json.dumps(value))
        else:
            set_parts.append(f"{key} = %s")
            params.append(value)
    params.append(job_id)
    query = f"""
        UPDATE import_jobs
        SET {', '.join(set_parts)}, updated_at = CURRENT_TIMESTAMP
        WHERE job_id = %s
    """
    cursor.execute(query, params)
    conn.commit()
    conn.close()

# Call table creation on startup
init_import_jobs_table()

# ---------- (license functions remain unchanged) ----------
def load_license_data():
    try:
        if os.path.exists(LICENSE_FILE):
            with open(LICENSE_FILE, "r") as f:
                return json.load(f)
    except Exception as e:
        print(f"Error loading license file: {e}")
    return {"trial_start": None, "licensed": False, "key": None}

def save_license_data(data):
    try:
        with open(LICENSE_FILE, "w") as f:
            json.dump(data, f, indent=2)
        return True
    except Exception as e:
        print(f"Error saving license file: {e}")
        return False

def get_trial_start():
    data = load_license_data()
    start = data.get("trial_start")
    if start:
        try:
            return datetime.fromisoformat(start)
        except:
            return None
    return None

def get_trial_end():
    start = get_trial_start()
    if start:
        return start + timedelta(days=TRIAL_DAYS)
    return None

def get_remaining_trial():
    data = load_license_data()
    if data.get("licensed", False):
        return None
    end = get_trial_end()
    if not end:
        return None
    now = datetime.now()
    if now >= end:
        return timedelta(0)
    return end - now

def is_trial_active():
    if load_license_data().get("licensed", False):
        return True
    rem = get_remaining_trial()
    return rem is not None and rem.total_seconds() > 0

# ---------------------- CONTEXT PROCESSOR ----------------------
@app.context_processor
def inject_user():
    if 'user_id' in session:
        return {
            'current_user': {
                'username': session.get('username'),
                'role': session.get('role')
            }
        }
    return {'current_user': None}

# ---------------------- LOGIN REQUIRED DECORATOR ----------------------
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Unauthorized', 'message': 'Please log in'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ---------------------- ADMIN REQUIRED DECORATOR ----------------------
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return render_template("error.html", message="Admin access required"), 403
        return f(*args, **kwargs)
    return decorated_function

# ---------------------- PUBLIC ROUTES ----------------------
@app.route("/login")
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template("login.html")

@app.route("/signup")
def signup():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return render_template("signup.html")

# ---------------------- PROTECTED ROUTES ----------------------
@app.route("/")
@login_required
def dashboard():
    return render_template("dashboard.html")

@app.route("/products")
@login_required
def products():
    return render_template("products.html")

@app.route("/products/screens")
@login_required
def products_screen():
    return render_template("products_screen.html")

@app.route("/products/add", methods=["POST"])
@login_required
def add_product():
    return redirect(url_for("products"))

@app.route("/sales")
@login_required
def sales():
    return render_template("sales.html")

@app.route("/sales/screens")
@login_required
def sales_screen():
    return render_template("sales_screen.html")

@app.route("/purchases")
@login_required
def purchases():
    return render_template("purchases.html")

@app.route("/purchases/screens")
@login_required
def purchases_screen():
    return render_template("purchases_screen.html")

@app.route("/analytics")
@login_required
def analytics():
    return render_template("analytics.html")

@app.route("/today-sales")
@login_required
def today_sales():
    return render_template("today_sales.html")

@app.route("/today-sales/screens")
@login_required
def today_sales_screen():
    return render_template("today_sales_screen.html")

@app.route("/archive")
@login_required
def archive():
    return render_template("archive.html")

# ---------------------- ADMIN PAGE ----------------------
@app.route("/admin/users")
@admin_required
def admin_users():
    return render_template("admin_users.html")

# ===================== LICENSE API =====================
@app.route('/api/license/status', methods=['GET'])
def api_license_status():
    data = load_license_data()
    if not data.get("trial_start"):
        data["trial_start"] = datetime.now().isoformat()
        save_license_data(data)
        data = load_license_data()
    rem = get_remaining_trial()
    remaining_seconds = int(rem.total_seconds()) if rem else 0
    return jsonify({
        "licensed": data.get("licensed", False),
        "trial_active": is_trial_active(),
        "remaining_seconds": remaining_seconds,
        "trial_days": TRIAL_DAYS
    })

@app.route('/api/license/activate', methods=['POST'])
def api_license_activate():
    req = request.json
    key = req.get("key", "").strip()
    if key == MASTER_KEY:
        data = load_license_data()
        data["licensed"] = True
        data["key"] = key
        save_license_data(data)
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "Invalid license key"}), 400

# ===================== AUTH API =====================
@app.route('/api/auth/login', methods=['POST'])
def api_auth_login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    user = login_user(username, password)
    if user:
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['role'] = user['role']
        return jsonify({'success': True, 'user': user})
    else:
        return jsonify({'success': False, 'error': 'Invalid credentials'}), 401

@app.route('/api/auth/signup', methods=['POST'])
def api_auth_signup():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    requested_role = data.get('role', 'user')
    
    current_admin_count = count_admins()
    
    if current_admin_count >= 2:
        role = 'user'
        print(f"⚠️ Already {current_admin_count} admins (max 2) - forcing role 'user' for {username}")
    else:
        if requested_role == 'admin':
            role = 'admin'
            print(f"✅ Admin {username} created. ({current_admin_count + 1}/2)")
        else:
            role = 'user'
            print(f"✅ Regular user {username} created.")
    
    success = create_user(username, password, role)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Username already exists'}), 400

@app.route('/api/auth/logout', methods=['POST'])
def api_auth_logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/check', methods=['GET'])
def api_auth_check():
    if 'user_id' in session:
        return jsonify({
            'logged_in': True,
            'username': session.get('username'),
            'role': session.get('role')
        })
    else:
        return jsonify({'logged_in': False})

@app.route('/api/auth/admin_exists', methods=['GET'])
def api_auth_admin_exists():
    exists = admin_exists()
    return jsonify({'admin_exists': exists})

@app.route('/api/auth/admin_count', methods=['GET'])
def api_auth_admin_count():
    count = count_admins()
    return jsonify({'admin_count': count})

# ===================== ADMIN USER MANAGEMENT API =====================
@app.route('/api/admin/users', methods=['GET'])
@admin_required
def api_admin_get_users():
    users = get_all_users()
    return jsonify(users)

@app.route('/api/admin/users', methods=['POST'])
@admin_required
def api_admin_create_user():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'user')
    if not username or not password:
        return jsonify({'success': False, 'error': 'Username and password required'}), 400
    success = create_user(username, password, role)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Username already exists'}), 400

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400
    success = delete_user(user_id)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Delete failed'}), 400

@app.route('/api/admin/users/<int:user_id>/role', methods=['PUT'])
@admin_required
def api_admin_update_role(user_id):
    data = request.json
    new_role = data.get('role')
    if new_role not in ['admin', 'user']:
        return jsonify({'success': False, 'error': 'Invalid role'}), 400
    success = update_user_role(user_id, new_role)
    if success:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Update failed'}), 400


# ===================== DASHBOARD API =====================
@app.route('/api/dashboard/summary', methods=['GET'])
@login_required
def api_dashboard_summary():
    selected_date_str = request.args.get('date')
    start_datetime = request.args.get('start_datetime')
    end_datetime = request.args.get('end_datetime')
    
    selected_date = None
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except:
            selected_date = None
    
    sales = get_today_sales(selected_date, start_datetime, end_datetime)
    profit = get_today_profit(selected_date, start_datetime, end_datetime)
    total_products = get_total_products()
    low_stock_products = get_low_stock_products(threshold=10)
    low_stock_count = len(low_stock_products)
    
    return jsonify({
        'sales': sales,
        'profit': profit,
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'low_stock_products': low_stock_products
    })

@app.route('/api/dashboard/top_products', methods=['GET'])
@login_required
def api_dashboard_top_products():
    selected_date_str = request.args.get('date')
    start_datetime = request.args.get('start_datetime')
    end_datetime = request.args.get('end_datetime')
    limit = int(request.args.get('limit', 10))
    
    selected_date = None
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except:
            selected_date = None
    
    top = get_top_products(selected_date, limit, start_datetime, end_datetime)
    result = [{'name': row[0], 'brand': row[1] or '', 'category': row[2] or '', 'qty': int(row[3])} for row in top]
    return jsonify(result)

@app.route('/api/dashboard/sales_history', methods=['GET'])
@login_required
def api_dashboard_sales_history():
    selected_date_str = request.args.get('date')
    start_datetime = request.args.get('start_datetime')
    end_datetime = request.args.get('end_datetime')
    
    selected_date = None
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except:
            selected_date = None
    
    history = get_sales_history(selected_date, start_datetime, end_datetime)
    result = []
    for row in history:
        date_str = row[0].isoformat() if hasattr(row[0], 'isoformat') else str(row[0])
        result.append({
            'date': date_str,
            'total_sales': float(row[1]),
            'profit': float(row[2]),
            'discount': float(row[3])
        })
    return jsonify(result)

# ===================== PURCHASES API (with category filter) =====================
def serialize_purchase(p):
    p_copy = p.copy()
    if 'date' in p_copy and p_copy['date']:
        if hasattr(p_copy['date'], 'isoformat'):
            p_copy['date'] = p_copy['date'].isoformat()
        else:
            p_copy['date'] = str(p_copy['date'])
    return p_copy

@app.route('/api/purchases', methods=['GET'])
@login_required
def api_get_purchases():
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    purchases = get_all_purchases()
    
    if category:
        purchases = [p for p in purchases if p.get('category') == category]
    if exclude_category:
        purchases = [p for p in purchases if p.get('category') != exclude_category]
    
    return jsonify([serialize_purchase(p) for p in purchases])

@app.route('/api/purchases', methods=['POST'])
@login_required
def api_add_purchase():
    data = request.json
    try:
        batch_id = add_purchase(
            name=data['name'],
            brand=data['brand'],
            category=data['category'],
            quantity=int(data['quantity']),
            cost_price=float(data['cost_price']),
            discount=float(data.get('discount', 0)),
            selling_price=float(data['selling_price'])
        )
        return jsonify({'success': True, 'batch_id': batch_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/purchases/<int:batch_id>', methods=['PUT'])
@login_required
def api_update_purchase(batch_id):
    data = request.json
    try:
        update_product(
            batch_id=batch_id,
            name=data['name'],
            brand=data['brand'],
            category=data['category'],
            quantity=int(data['quantity']),
            cost_price=float(data['cost_price']),
            discount=float(data.get('discount', 0)),
            selling_price=float(data['selling_price'])
        )
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/purchases/filter', methods=['GET'])
@login_required
def api_filter_purchases():
    start = request.args.get('start_date')
    end = request.args.get('end_date')
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    if not start or not end:
        return jsonify([])
    purchases = get_purchases_by_date_range(start, end)
    if category:
        purchases = [p for p in purchases if p.get('category') == category]
    if exclude_category:
        purchases = [p for p in purchases if p.get('category') != exclude_category]
    return jsonify([serialize_purchase(p) for p in purchases])

@app.route('/api/purchases/suggestions/name', methods=['GET'])
@login_required
def api_suggest_name():
    q = request.args.get('q', '')
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    if not q:
        return jsonify([])
    suggestions = get_product_suggestions(q)
    if category:
        suggestions = [s for s in suggestions if s.get('category') == category]
    if exclude_category:
        suggestions = [s for s in suggestions if s.get('category') != exclude_category]
    return jsonify(suggestions)

@app.route('/api/purchases/suggestions/category', methods=['GET'])
@login_required
def api_suggest_category():
    q = request.args.get('q', '')
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    if not q:
        return jsonify([])
    suggestions = get_category_suggestions(q)
    if category:
        suggestions = [s for s in suggestions if s.get('category') == category]
    if exclude_category:
        suggestions = [s for s in suggestions if s.get('category') != exclude_category]
    return jsonify(suggestions)

@app.route('/api/purchases/pdf', methods=['POST'])
@login_required
def api_purchases_pdf():
    data = request.json
    purchases = data.get('purchases', [])
    from_date = data.get('from_date', '')
    to_date = data.get('to_date', '')
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=15, leftMargin=15,
                            topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("📦 Purchases Report", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    if from_date and to_date:
        elements.append(Paragraph(f"Date Range: {from_date} → {to_date}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    table_data = [["ID", "Name", "Brand", "Qty", "Stock", "Cost", "Discount", "Total", "Selling", "Date/Time"]]
    total_qty = total_cost = total_discount = total_selling = 0
    row_colors = [colors.whitesmoke, colors.lightgrey]

    for p in purchases:
        date_str = p.get('date', '')
        if date_str:
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                date_str = dt.strftime("%Y-%m-%d %H:%M")
            except:
                date_str = str(date_str)[:16]

        table_data.append([
            p["batch_id"], p["name"], p["brand"],
            p["quantity"], p["remaining_quantity"],
            f"₵{p.get('cost_price', 0):.2f}",
            f"₵{p.get('discount', 0):.2f}",
            f"₵{p.get('total_cost', 0):.2f}",
            f"₵{p.get('selling_price', 0):.2f}",
            date_str
        ])
        total_qty += p.get("quantity", 0)
        total_cost += p.get("total_cost", 0)
        total_discount += p.get("discount", 0)
        total_selling += p.get("selling_price", 0) * p.get("quantity", 0)

    table = Table(table_data, repeatRows=1, hAlign='LEFT',
                  colWidths=[1.8*cm, 4.5*cm, 4*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3.5*cm])
    style = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#00CFCF")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("ALIGN", (3,1), (-2,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.black),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
    ])
    for i in range(1, len(table_data)):
        style.add("BACKGROUND", (0,i), (-1,i), row_colors[i%2])
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1,12))

    summary = f"Total Qty: {total_qty} | Total Discount: ₵{total_discount:.2f} | Total Cost: ₵{total_cost:.2f} | Total Selling: ₵{total_selling:.2f}"
    elements.append(Paragraph(summary, styles["Heading2"]))
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"Purchases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                     mimetype='application/pdf')

# ===================== PRODUCT API =====================
@app.route('/api/products', methods=['GET'])
@login_required
def api_get_products():
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    purchases = get_all_purchases()
    all_products = get_all_products_service()
    id_map = {(prod['name'], prod['brand']): prod['product_id'] for prod in all_products}
    products_dict = {}
    total_batches = 0
    for p in purchases:
        total_batches += 1
        key = (p['name'], p['brand'])
        if key not in products_dict:
            prod_category = p.get('category')
            if category and prod_category != category:
                continue
            if exclude_category and prod_category == exclude_category:
                continue
            products_dict[key] = {
                'name': p['name'],
                'brand': p['brand'],
                'category': prod_category,
                'cost_price': p['cost_price'],
                'selling_price': p['selling_price'],
                'discount': p['discount'],
                'stock': 0,
                'batches': [],
                'product_id': id_map.get(key)
            }
        if key in products_dict:
            products_dict[key]['stock'] += p['remaining_quantity']
            products_dict[key]['batches'].append({
                'batch_id': p['batch_id'],
                'quantity': p['quantity'],
                'remaining_quantity': p['remaining_quantity'],
                'cost_price': p['cost_price'],
                'selling_price': p['selling_price'],
                'discount': p['discount'],
                'date': p['date']
            })
    result = list(products_dict.values())
    return jsonify({
        'products': result,
        'total_products': len(result),
        'total_batches': total_batches
    })

@app.route('/api/products', methods=['POST'])
@login_required
def api_add_product():
    data = request.json
    try:
        batch_id = add_purchase(
            name=data['name'],
            brand=data['brand'],
            category=data.get('category', ''),
            quantity=int(data['quantity']),
            cost_price=float(data['cost_price']),
            discount=float(data.get('discount', 0)),
            selling_price=float(data['selling_price'])
        )
        return jsonify({'success': True, 'batch_id': batch_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/products/<int:product_id>', methods=['DELETE'])
@login_required
def api_delete_product(product_id):
    delete_type = request.args.get('type', 'keep')
    try:
        if delete_type == 'clean':
            delete_product_clean_all(product_id)
        else:
            delete_product_keep_history(product_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/batches/<int:batch_id>', methods=['DELETE'])
@login_required
def api_delete_batch(batch_id):
    delete_type = request.args.get('type', 'keep')
    try:
        if delete_type == 'clean':
            delete_batch_clean_all(batch_id)
        else:
            delete_batch(batch_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# ===================== SALES API =====================
@app.route('/api/sales/products', methods=['GET'])
@login_required
def api_sales_products():
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    products = get_products_for_sale()
    if category:
        products = [p for p in products if p.get('category') == category]
    if exclude_category:
        products = [p for p in products if p.get('category') != exclude_category]
    return jsonify(products)

@app.route('/api/sales/batches/<int:product_id>', methods=['GET'])
@login_required
def api_sales_batches(product_id):
    batches = get_batches_for_product(product_id)
    return jsonify(batches)

@app.route('/api/sales/complete', methods=['POST'])
@login_required
def api_sales_complete():
    data = request.json
    cart_items = data.get('cart_items', [])
    sale_datetime = data.get('sale_datetime')
    selected_batches = data.get('selected_batches', [])
    payment_method = data.get('payment_method', 'cash')
    cheque_number = data.get('cheque_number')
    
    try:
        result = create_multi_sale(
            cart_items, 
            sale_datetime, 
            selected_batches,
            payment_method,
            cheque_number
        )
        
        receipt_cart = []
        for idx, item in enumerate(cart_items):
            product_batches = []
            for sb in selected_batches:
                if sb.get('product_id') == item['product']['id']:
                    batch_info = get_batch_by_id(sb['batch_id'])
                    if batch_info:
                        product_batches.append({
                            'batch_id': sb['batch_id'],
                            'qty': sb['qty'],
                            'batch': batch_info
                        })
            receipt_cart.append({
                'product': item['product'],
                'qty': item['qty'],
                'discount': item['discount'],
                'selected_batches': product_batches
            })
        
        receipt_file, receipt_text = generate_receipt_multi(
            receipt_cart, 
            result['total'],
            payment_method,
            cheque_number
        )
        
        return jsonify({
            'success': True,
            'sale_id': result['sale_id'],
            'total': result['total'],
            'date': result['date'],
            'receipt_text': receipt_text,
            'receipt_file': receipt_file,
            'payment_method': payment_method,
            'cheque_number': cheque_number
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

# ===================== REVERSE SALE API =====================
@app.route('/api/sales/reverse/<int:sale_id>', methods=['POST'])
@login_required
def api_reverse_sale(sale_id):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT reversed FROM sales WHERE id = %s", (sale_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Sale not found'}), 404
        if row[0] == 1:
            return jsonify({'success': False, 'error': 'Sale already reversed'}), 400
        cursor.execute("""
            SELECT product_id, batch_id, quantity, cost_price, selling_price, profit
            FROM sales_items
            WHERE sale_id = %s
        """, (sale_id,))
        items = cursor.fetchall()
        if not items:
            return jsonify({'success': False, 'error': 'No items found for this sale'}), 400
        for item in items:
            product_id, batch_id, qty, cost_price, selling_price, profit = item
            cursor.execute("""
                UPDATE purchase_batches
                SET remaining_quantity = remaining_quantity + %s
                WHERE id = %s
            """, (qty, batch_id))
            update_product_stock(cursor, product_id)
        cursor.execute("UPDATE sales SET reversed = 1 WHERE id = %s", (sale_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/sales/reverse_items', methods=['POST'])
@login_required
def api_reverse_sale_items():
    import traceback
    data = request.json
    sale_id = data.get('sale_id')
    items_to_reverse = data.get('items', [])
    if not sale_id or not items_to_reverse:
        return jsonify({'success': False, 'error': 'Missing sale_id or items'}), 400
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("SELECT reversed, subtotal, discount, total, profit FROM sales WHERE id = %s", (sale_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': False, 'error': 'Sale not found'}), 404
        if row[0] == 1:
            return jsonify({'success': False, 'error': 'Sale already fully reversed'}), 400
        original_subtotal = row[1]
        original_discount = row[2]
        original_net_profit = row[4]
        reversed_subtotal = 0
        reversed_profit_gross = 0
        for item in items_to_reverse:
            batch_id = item['batch_id']
            qty = item['quantity']
            cursor.execute("SELECT product_id, selling_price FROM purchase_batches WHERE id = %s", (batch_id,))
            batch_row = cursor.fetchone()
            if not batch_row:
                raise ValueError(f"Batch {batch_id} not found")
            product_id = batch_row[0]
            selling_price = batch_row[1]
            cursor.execute("SELECT cost_price, profit FROM sales_items WHERE sale_id = %s AND batch_id = %s", (sale_id, batch_id))
            item_row = cursor.fetchone()
            if not item_row:
                raise ValueError(f"Sales item for batch {batch_id} not found")
            cost_price = item_row[0]
            item_gross_profit = item_row[1]
            item_subtotal = selling_price * qty
            reversed_subtotal += item_subtotal
            reversed_profit_gross += item_gross_profit
            cursor.execute("UPDATE purchase_batches SET remaining_quantity = remaining_quantity + %s WHERE id = %s", (qty, batch_id))
            cursor.execute("DELETE FROM sales_items WHERE sale_id = %s AND batch_id = %s", (sale_id, batch_id))
            update_product_stock(cursor, product_id)
        new_subtotal = original_subtotal - reversed_subtotal
        new_gross_profit = original_net_profit + original_discount - reversed_profit_gross
        if original_subtotal > 0:
            new_discount = original_discount * (new_subtotal / original_subtotal)
        else:
            new_discount = 0
        new_net_profit = new_gross_profit - new_discount
        new_total = new_subtotal - new_discount
        if new_total < 0:
            new_total = 0
        if new_subtotal == 0:
            cursor.execute("UPDATE sales SET reversed = 1 WHERE id = %s", (sale_id,))
        else:
            cursor.execute("""
                UPDATE sales
                SET subtotal = %s, discount = %s, total = %s, profit = %s
                WHERE id = %s
            """, (new_subtotal, new_discount, new_total, new_net_profit, sale_id))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        print("ERROR in reverse_items:")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

# ===================== TODAY'S SALES API (with category filter) =====================
@app.route('/api/today_sales', methods=['GET'])
@login_required
def api_today_sales():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    exclude_category = request.args.get('exclude_category')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    select_clause = """
        SELECT 
            sales.id, 
            products.name, 
            products.brand, 
            products.category,
            sales_items.quantity, 
            sales_items.selling_price,
            sales_items.quantity * sales_items.selling_price AS subtotal,
            sales.discount, 
            sales.total, 
            sales_items.profit,
            COALESCE(purchase_batches.id, -1) as batch_id,
            COALESCE(purchase_batches.cost_price, 0) as cost_price, 
            sales.date,
            CASE WHEN purchase_batches.id IS NULL THEN 1 ELSE 0 END as is_deleted_batch,
            sales.profit as net_profit,
            COALESCE(sales.payment_method, 'cash') as payment_method,
            sales.cheque_number
    """
    from_clause = """
        FROM sales
        JOIN sales_items ON sales.id = sales_items.sale_id
        JOIN products ON products.id = sales_items.product_id
        LEFT JOIN purchase_batches ON purchase_batches.id = sales_items.batch_id
    """
    
    where_conditions = ["sales.reversed = 0"]
    if start_date and end_date:
        where_conditions.append("sales.date::date BETWEEN %s AND %s")
    if category:
        where_conditions.append("products.category = %s")
    if exclude_category:
        where_conditions.append("products.category != %s")
    
    where_clause = " AND ".join(where_conditions)
    query = f"{select_clause} {from_clause} WHERE {where_clause} ORDER BY sales.date DESC"
    
    params = []
    if start_date and end_date:
        params.extend([start_date, end_date])
    if category:
        params.append(category)
    if exclude_category:
        params.append(exclude_category)
    
    try:
        cursor.execute(query, params)
        rows = cursor.fetchall()
        sales_data = []
        for r in rows:
            sales_data.append({
                'sale_id': r[0],
                'name': r[1],
                'brand': r[2] or '',
                'category': r[3] or '',
                'quantity': r[4],
                'selling_price': float(r[5]),
                'subtotal': float(r[6]),
                'discount': float(r[7]),
                'total': float(r[8]),
                'profit': float(r[9]),
                'batch_id': r[10],
                'cost_price': float(r[11]),
                'sale_date': r[12].isoformat() if hasattr(r[12], 'isoformat') else str(r[12]),
                'is_deleted_batch': bool(r[13]),
                'net_profit': float(r[14]),
                'payment_method': r[15] if len(r) > 15 else 'cash',
                'cheque_number': r[16] if len(r) > 16 else None
            })
        return jsonify(sales_data)
    except Exception as e:
        print(f"Error in today_sales API: {str(e)}")
        return jsonify([]), 500
    finally:
        conn.close()

@app.route('/api/today_sales/pdf', methods=['POST'])
@login_required
def api_today_sales_pdf():
    data = request.json
    sales_data = data.get('sales_data', [])
    period_text = data.get('period_text', 'Sales Report')
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Sales Report", styles['Title']))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(period_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*cm))
    seen_sales = set()
    total_sales = 0.0
    total_discount = 0.0
    total_profit = 0.0
    total_items = 0
    for sale in sales_data:
        sale_id = sale['sale_id']
        if sale_id not in seen_sales:
            total_sales += sale['total']
            total_discount += sale['discount']
            seen_sales.add(sale_id)
        total_profit += sale['profit']
        total_items += sale['quantity']
    elements.append(Paragraph(
        f"🧾 Items: {total_items}   |   💰 Sales: ₵{total_sales:.2f}   |   "
        f"📉 Discount: ₵{total_discount:.2f}   |   📈 Profit: ₵{total_profit:.2f}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3*cm))
    table_data = [["Name", "Brand", "Category", "Qty", "Price", "Subtotal", 
                   "Discount", "Total", "Profit", "Batch", "Cost", "Sale Date", "Payment", "Status"]]
    for s in sales_data:
        status = "Deleted Batch" if s['is_deleted_batch'] else "Active"
        payment_display = s.get('payment_method', 'cash').upper()
        if s.get('cheque_number'):
            payment_display += f" #{s['cheque_number']}"
        row = [
            s['name'], s['brand'], s['category'],
            str(s['quantity']),
            f"₵{s['selling_price']:.2f}",
            f"₵{s['subtotal']:.2f}",
            f"₵{s['discount']:.2f}",
            f"₵{s['total']:.2f}",
            f"₵{s['profit']:.2f}",
            str(s['batch_id']) if s['batch_id'] != -1 else "DELETED",
            f"₵{s['cost_price']:.2f}",
            s['sale_date'],
            payment_display,
            status
        ]
        table_data.append(row)
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (3,1), (-1,-1), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"SalesReport_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                     mimetype='application/pdf')

# ===================== ANALYTICS API =====================
@app.route('/api/analytics/summary', methods=['GET'])
@login_required
def api_analytics_summary():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    if start_date and end_date:
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE s.date::date BETWEEN %s AND %s
            AND s.reversed = 0
        """, (start_date, end_date))
    elif period == 'weekly':
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE s.date >= CURRENT_DATE - INTERVAL '6 days'
            AND s.reversed = 0
        """)
    elif period == 'monthly':
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND EXTRACT(MONTH FROM s.date) = EXTRACT(MONTH FROM CURRENT_DATE)
            AND s.reversed = 0
        """)
    elif period == 'yearly':
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND s.reversed = 0
        """)
    elif period == 'all':
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE s.reversed = 0
        """)
    else:
        cursor.execute("""
            SELECT 
                COALESCE(SUM(s.total), 0),
                COALESCE(SUM(s.discount), 0),
                COALESCE(SUM(s.profit), 0),
                COALESCE(SUM(si.quantity), 0)
            FROM sales s
            LEFT JOIN sales_items si ON s.id = si.sale_id
            WHERE s.date::date = CURRENT_DATE
            AND s.reversed = 0
        """)
    
    result = cursor.fetchone()
    conn.close()
    
    return jsonify({
        'items_sold': int(result[3] or 0),
        'subtotal': float(result[0] or 0),
        'discount': float(result[1] or 0),
        'total': float(result[0] or 0),
        'profit': float(result[2] or 0)
    })

@app.route('/api/analytics/trend', methods=['GET'])
@login_required
def api_analytics_trend():
    period = request.args.get('period', 'daily')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    if start_date and end_date:
        cursor.execute("""
            SELECT 
                s.date::date as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE s.date::date BETWEEN %s AND %s
            AND s.reversed = 0
            GROUP BY s.date::date
            ORDER BY label ASC
        """, (start_date, end_date))
    elif period == 'weekly':
        cursor.execute("""
            SELECT 
                s.date::date as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE s.date >= CURRENT_DATE - INTERVAL '6 days'
            AND s.reversed = 0
            GROUP BY s.date::date
            ORDER BY label ASC
        """)
    elif period == 'monthly':
        cursor.execute("""
            SELECT 
                s.date::date as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND EXTRACT(MONTH FROM s.date) = EXTRACT(MONTH FROM CURRENT_DATE)
            AND s.reversed = 0
            GROUP BY s.date::date
            ORDER BY label ASC
        """)
    elif period == 'yearly':
        cursor.execute("""
            SELECT 
                DATE_TRUNC('month', s.date) as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND s.reversed = 0
            GROUP BY DATE_TRUNC('month', s.date)
            ORDER BY label ASC
        """)
    elif period == 'all':
        cursor.execute("""
            SELECT 
                DATE_TRUNC('month', s.date) as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE s.reversed = 0
            GROUP BY DATE_TRUNC('month', s.date)
            ORDER BY label ASC
            LIMIT 24
        """)
    else:
        cursor.execute("""
            SELECT 
                s.date::date as label,
                COALESCE(SUM(s.total), 0) as sales_total,
                COALESCE(SUM(s.profit), 0) as profit_total
            FROM sales s
            WHERE s.date::date = CURRENT_DATE
            AND s.reversed = 0
            GROUP BY s.date::date
            ORDER BY label ASC
        """)
    
    rows = cursor.fetchall()
    conn.close()
    
    result = []
    for row in rows:
        label = row[0]
        if period in ['yearly', 'all']:
            if hasattr(label, 'strftime'):
                label = label.strftime('%b %Y')
            else:
                label = str(label)
        else:
            if hasattr(label, 'strftime'):
                label = label.strftime('%Y-%m-%d')
            else:
                label = str(label)
        
        result.append({
            'label': str(label),
            'sales': float(row[1]),
            'profit': float(row[2])
        })
    
    return jsonify(result)

@app.route('/api/analytics/top_products', methods=['GET'])
@login_required
def api_analytics_top_products():
    period = request.args.get('period', 'daily')
    limit = int(request.args.get('limit', 10))
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    conn = get_connection()
    cursor = conn.cursor()
    
    if start_date and end_date:
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.date::date BETWEEN %s AND %s
            AND s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (start_date, end_date, limit))
    elif period == 'weekly':
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.date >= CURRENT_DATE - INTERVAL '6 days'
            AND s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (limit,))
    elif period == 'monthly':
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND EXTRACT(MONTH FROM s.date) = EXTRACT(MONTH FROM CURRENT_DATE)
            AND s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (limit,))
    elif period == 'yearly':
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE EXTRACT(YEAR FROM s.date) = EXTRACT(YEAR FROM CURRENT_DATE)
            AND s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (limit,))
    elif period == 'all':
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (limit,))
    else:
        cursor.execute("""
            SELECT 
                p.name, 
                COALESCE(p.brand, '') as brand, 
                COALESCE(p.category, '') as category, 
                COALESCE(SUM(si.quantity), 0) as qty
            FROM sales_items si
            JOIN sales s ON si.sale_id = s.id
            JOIN products p ON p.id = si.product_id
            WHERE s.date::date = CURRENT_DATE
            AND s.reversed = 0
            AND NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = p.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
            GROUP BY si.product_id, p.name, p.brand, p.category
            ORDER BY qty DESC
            LIMIT %s
        """, (limit,))
    
    rows = cursor.fetchall()
    conn.close()
    
    result = [{'name': r[0], 'brand': r[1] or '', 'category': r[2] or '', 'quantity': int(r[3])} for r in rows]
    return jsonify(result)

# ===================== ARCHIVE API =====================
@app.route('/api/archive', methods=['GET'])
@login_required
def api_archive():
    status_filter = request.args.get('status', 'ALL')
    print(f"📁 Archive API called with status filter: {status_filter}")
    
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT 
                id,
                name,
                brand,
                category,
                cost_price,
                selling_price,
                stock,
                discount
            FROM products
            WHERE NOT EXISTS (
                SELECT 1 FROM deleted_products dp 
                WHERE dp.product_id = products.id 
                AND dp.action = 'PERMANENTLY DELETED' 
                AND dp.source = 'product'
            )
        """)
        active_rows = cursor.fetchall()
        print(f"✅ Found {len(active_rows)} active products")
        
        active_items = []
        for r in active_rows:
            active_items.append({
                'id': None,
                'name': r[1],
                'brand': r[2] or '',
                'category': r[3] or '',
                'cost': float(r[4] or 0),
                'price': float(r[5] or 0),
                'stock': int(r[6] or 0),
                'discount': float(r[7] or 0),
                'action': 'ACTIVE',
                'date': '',
                'source': 'active',
                'is_permanent': False,
                'batch_id': None,
                'batch_quantity': None,
                'batch_remaining': None,
                'product_id': r[0]
            })
        
        cursor.execute("""
            SELECT 
                id, name, brand, cost_price, selling_price, stock, 
                category, discount, action, deleted_at, batch_id, 
                batch_quantity, batch_remaining, product_id, source
            FROM deleted_products
            ORDER BY deleted_at DESC
        """)
        deleted_rows = cursor.fetchall()
        print(f"✅ Found {len(deleted_rows)} deleted records")
        
        deleted_items = []
        for r in deleted_rows:
            action = str(r[8]).upper() if len(r) > 8 and r[8] else 'UNKNOWN'
            is_permanent = action == 'PERMANENTLY DELETED'
            deleted_items.append({
                'id': r[0],
                'name': r[1] if r[1] else '-',
                'brand': r[2] if r[2] else '-',
                'category': r[6] if r[6] else '-',
                'cost': float(r[3] or 0),
                'price': float(r[4] or 0),
                'stock': int(r[5] or 0),
                'discount': float(r[7] or 0),
                'action': action,
                'date': r[9].isoformat() if hasattr(r[9], 'isoformat') else str(r[9]) if r[9] else '',
                'source': r[14] if len(r) > 14 and r[14] else 'product',
                'is_permanent': is_permanent,
                'batch_id': r[10] if len(r) > 10 else None,
                'batch_quantity': r[11] if len(r) > 11 else None,
                'batch_remaining': r[12] if len(r) > 12 else None,
                'product_id': r[13] if len(r) > 13 else None
            })
        
        combined = active_items + deleted_items
        
        if status_filter != 'ALL':
            combined = [item for item in combined if item['action'] == status_filter]
            print(f"📊 Filtered to {len(combined)} items with status '{status_filter}'")
        
        def sort_key(item):
            if item['action'] == 'ACTIVE':
                return (datetime.max, item['name'])
            else:
                try:
                    date_obj = datetime.fromisoformat(str(item['date'])) if item['date'] else datetime.min
                except:
                    date_obj = datetime.min
                return (date_obj, item['name'])
        
        combined.sort(key=sort_key, reverse=True)
        
        print(f"📦 Returning {len(combined)} items")
        return jsonify(combined)
        
    except Exception as e:
        print(f"❌ Error in archive API: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/archive/restore', methods=['POST'])
@login_required
def api_archive_restore():
    data = request.json
    archive_id = data.get('archive_id')
    if not archive_id:
        return jsonify({'success': False, 'error': 'Missing archive_id'}), 400
    try:
        restore_archive(archive_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/archive/batches', methods=['GET'])
@login_required
def api_archive_batches():
    name = request.args.get('name')
    brand = request.args.get('brand')
    if not name or not brand:
        return jsonify([])
    purchases = get_all_purchases()
    batches = [b for b in purchases if b['name'] == name and b['brand'] == brand]
    return jsonify(batches)

# ===================== IMPORT ENDPOINTS (with progress) =====================
import openpyxl
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import psycopg2
from psycopg2 import sql

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/inventory/import', methods=['GET'])
@admin_required
def import_inventory_page():
    return render_template('import_inventory.html')

# ----- INVENTORY IMPORT (background thread) with direct connection -----
def run_inventory_import(job_id, file_stream, target_category):
    try:
        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active
    except Exception as e:
        update_job_progress(job_id, status='error', errors=[f"Unable to read workbook: {str(e)}"])
        return

    # Find header row
    header_row_idx = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        if row and any(cell and isinstance(cell, str) and ('item' in cell.lower() or 'qty' in cell.lower() or 'rate' in cell.lower()) for cell in row):
            header_row_idx = i + 1
            header_row = row
            break

    if header_row_idx is None:
        update_job_progress(job_id, status='error', errors=['Could not find header row'])
        return

    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell:
            cell_lower = str(cell).strip().lower()
            if cell_lower in ['item', 'product', 'name', 'details']:
                header_map['name'] = idx
            elif cell_lower in ['qty', 'quantity']:
                header_map['quantity'] = idx
            elif cell_lower in ['rate', 'cost', 'cost price', 'unit cost']:
                header_map['cost_price'] = idx
            elif cell_lower in ['date', 'purchase date']:
                header_map['date'] = idx
            elif cell_lower in ['discount']:
                header_map['discount'] = idx

    required = ['name']
    missing = [f for f in required if f not in header_map]
    if missing:
        update_job_progress(job_id, status='error', errors=[f'Missing column: {", ".join(missing)} (only Name is required)'])
        return

    rows_to_process = []
    skipped_rows = []   # only rows with no product name or fatal errors
    warning_rows = []   # rows with missing numeric fields (will be imported with defaults)

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not any(row):
            continue
        try:
            # ---- Product name ----
            name = str(row[header_map['name']]).strip() if row[header_map['name']] else ''
            if not name:
                skipped_rows.append({
                    'row': row_idx,
                    'data': {
                        'name': None,
                        'qty': row[header_map.get('quantity')] if header_map.get('quantity') is not None else None,
                        'rate': row[header_map.get('cost_price')] if header_map.get('cost_price') is not None else None
                    },
                    'reason': 'Product name is empty (row skipped)'
                })
                continue

            # ---- Quantity (allow 0 or missing) ----
            try:
                quantity = float(row[header_map['quantity']]) if header_map.get('quantity') is not None and row[header_map['quantity']] is not None else 0
                if quantity < 0:
                    quantity = 0
                    warning_rows.append(f"Row {row_idx}: Quantity was negative, set to 0")
            except:
                quantity = 0
                warning_rows.append(f"Row {row_idx}: Invalid quantity, set to 0")

            # ---- Cost price (allow 0 or missing) ----
            try:
                cost_price = float(row[header_map['cost_price']]) if header_map.get('cost_price') is not None and row[header_map['cost_price']] is not None else 0
                if cost_price < 0:
                    cost_price = 0
                    warning_rows.append(f"Row {row_idx}: Cost price was negative, set to 0")
            except:
                cost_price = 0
                warning_rows.append(f"Row {row_idx}: Invalid cost price, set to 0")

            # ---- Discount ----
            try:
                discount = float(row[header_map.get('discount')]) if header_map.get('discount') is not None and row[header_map['discount']] is not None else 0
                if discount < 0:
                    discount = 0
            except:
                discount = 0

            # ---- Selling price ----
            # If missing, set to 0 (user will update later)
            try:
                selling_price = float(row[header_map.get('selling_price')]) if header_map.get('selling_price') is not None and row[header_map['selling_price']] is not None else 0
                if selling_price < 0:
                    selling_price = 0
            except:
                selling_price = 0

            # ---- Purchase date ----
            purchase_date = None
            if 'date' in header_map and row[header_map['date']] is not None:
                try:
                    if isinstance(row[header_map['date']], (int, float)):
                        purchase_date = datetime(1899, 12, 30) + timedelta(days=row[header_map['date']])
                    else:
                        date_str = str(row[header_map['date']]).strip()
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                            try:
                                purchase_date = datetime.strptime(date_str, fmt)
                                break
                            except:
                                continue
                        if purchase_date is None:
                            purchase_date = datetime.now()
                except:
                    purchase_date = datetime.now()
            else:
                purchase_date = datetime.now()

            rows_to_process.append({
                'row_idx': row_idx,
                'name': name,
                'quantity': int(quantity),
                'cost_price': cost_price,
                'discount': discount,
                'selling_price': selling_price,
                'category': target_category,
                'purchase_date': purchase_date
            })
        except Exception as e:
            skipped_rows.append({
                'row': row_idx,
                'data': {
                    'name': row[header_map.get('name')] if header_map.get('name') is not None else None,
                    'qty': row[header_map.get('quantity')] if header_map.get('quantity') is not None else None,
                    'rate': row[header_map.get('cost_price')] if header_map.get('cost_price') is not None else None
                },
                'reason': f'Fatal error: {str(e)}'
            })

    total_rows = len(rows_to_process)
    update_job_progress(job_id, 
                        total=total_rows, 
                        processed=0, 
                        status='processing',
                        errors=warning_rows,  # warnings are stored in errors (not critical)
                        result={'imported': 0, 'skipped': skipped_rows, 'warnings': warning_rows, 'message': 'Parsing completed'})

    if total_rows == 0:
        update_job_progress(job_id, status='done',
                            result={'imported': 0, 'skipped': skipped_rows, 'warnings': warning_rows, 'message': 'No valid rows to import (all rows missing product name or had fatal errors)'})
        return

    BATCH_SIZE = 5000
    imported_count = 0
    overall_errors = warning_rows.copy()

    if not DATABASE_URL:
        update_job_progress(job_id, status='error', errors=['DATABASE_URL not configured'])
        return

    for i in range(0, total_rows, BATCH_SIZE):
        batch = rows_to_process[i:i+BATCH_SIZE]
        conn = None
        try:
            conn = psycopg2.connect(DATABASE_URL, sslmode='require')
            conn.autocommit = False
            cursor = conn.cursor()

            for item in batch:
                # Product lookup (ignores category, updates if needed)
                cursor.execute(
                    "SELECT id, category FROM products WHERE name = %s AND brand = ''",
                    (item['name'],)
                )
                result = cursor.fetchone()
                if result:
                    product_id = result[0]
                    existing_category = result[1]
                    if existing_category != item['category']:
                        cursor.execute(
                            "UPDATE products SET category = %s WHERE id = %s",
                            (item['category'], product_id)
                        )
                else:
                    cursor.execute(
                        "INSERT INTO products (name, brand, category, cost_price, selling_price, discount, stock) VALUES (%s, %s, %s, %s, %s, %s, 0) RETURNING id",
                        (item['name'], '', item['category'], item['cost_price'], item['selling_price'], item['discount'])
                    )
                    product_id = cursor.fetchone()[0]

                # Insert purchase batch
                cursor.execute("""
                    INSERT INTO purchase_batches
                    (product_id, quantity, remaining_quantity, cost_price, selling_price, discount, date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (product_id, item['quantity'], item['quantity'], item['cost_price'], item['selling_price'], item['discount'], item['purchase_date']))

                # Update product stock (even if quantity is 0, stock remains unchanged)
                if item['quantity'] != 0:
                    cursor.execute("""
                        UPDATE products
                        SET stock = stock + %s
                        WHERE id = %s
                    """, (item['quantity'], product_id))
                else:
                    # If quantity is 0, we still ensure stock is recalculated (optional)
                    # For simplicity, we'll just leave stock as is (it should already be 0 for new products)
                    pass

                imported_count += 1
                if imported_count % 50 == 0:
                    update_job_progress(job_id, processed=imported_count)

            conn.commit()
            update_job_progress(job_id, processed=imported_count)

        except Exception as e:
            if conn:
                conn.rollback()
            overall_errors.append(f"Batch starting at row {batch[0]['row_idx']}: {str(e)}")
            update_job_progress(job_id, errors=overall_errors)
            break
        finally:
            if conn:
                conn.close()

    if overall_errors and len(overall_errors) > 0:
        update_job_progress(job_id, status='error', errors=overall_errors)
    else:
        update_job_progress(job_id, status='done',
                            result={
                                'imported': imported_count,
                                'skipped': skipped_rows,
                                'warnings': warning_rows,
                                'message': f'Imported {imported_count} records, {len(skipped_rows)} rows skipped (missing product name), {len(warning_rows)} warnings (defaults applied)'
                            })

@app.route('/api/inventory/import', methods=['POST'])
@admin_required
def api_import_inventory():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'File type not allowed'}), 400

    target_category = request.form.get('target_category', 'Accessory')
    if target_category not in ['Accessory', 'Screen']:
        return jsonify({'success': False, 'error': 'Invalid target category'}), 400

    file_content = file.read()
    file_stream = io.BytesIO(file_content)

    job_id = str(uuid.uuid4())

    # Insert initial job record
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO import_jobs (job_id, status, total, processed, errors, result)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (job_id, 'pending', 0, 0, '[]', None))
    conn.commit()
    conn.close()

    thread = threading.Thread(
        target=run_inventory_import,
        args=(job_id, file_stream, target_category)
    )
    thread.daemon = True
    thread.start()

    return jsonify({'success': True, 'job_id': job_id})
# ----- SALES IMPORT (background thread) with direct connection -----
def run_sales_import(job_id, file_stream, target_category):
    try:
        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active
    except Exception as e:
        update_job_progress(job_id, status='error', errors=[f"Unable to read workbook: {str(e)}"])
        return

    # Find header
    header_row_idx = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        if row and any(cell and isinstance(cell, str) and 
                       ('item' in cell.lower() or 'qty' in cell.lower() or 'rate' in cell.lower()) for cell in row):
            header_row_idx = i + 1
            header_row = row
            break

    if header_row_idx is None:
        update_job_progress(job_id, status='error', errors=['Could not find header row'])
        return

    header_map = {}
    for idx, cell in enumerate(header_row):
        if cell:
            cell_lower = str(cell).strip().lower()
            if cell_lower in ['item', 'product', 'name', 'details']:
                header_map['item'] = idx
            elif cell_lower in ['qty', 'quantity']:
                header_map['qty'] = idx
            elif cell_lower in ['rate', 'selling price', 'unit price']:
                header_map['rate'] = idx
            elif cell_lower in ['date', 'sale date']:
                header_map['date'] = idx
            elif cell_lower in ['discount']:
                header_map['discount'] = idx

    required = ['item', 'qty', 'rate']
    missing = [f for f in required if f not in header_map]
    if missing:
        update_job_progress(job_id, status='error', errors=[f'Missing columns: {", ".join(missing)}'])
        return

    rows_to_process = []
    skipped_rows = []
    error_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if not any(row):
            continue
        try:
            item = str(row[header_map['item']]).strip() if row[header_map['item']] else ''
            if not item:
                skipped_rows.append({
                    'row': row_idx,
                    'data': {
                        'item': None,
                        'qty': row[header_map.get('qty')] if header_map.get('qty') is not None else None,
                        'rate': row[header_map.get('rate')] if header_map.get('rate') is not None else None
                    },
                    'reason': 'Product name is empty'
                })
                continue

            qty = float(row[header_map['qty']]) if row[header_map['qty']] is not None else 0
            if qty <= 0:
                skipped_rows.append({
                    'row': row_idx,
                    'data': {
                        'item': item,
                        'qty': qty,
                        'rate': row[header_map.get('rate')] if header_map.get('rate') is not None else None
                    },
                    'reason': f'Quantity must be positive (got {qty})'
                })
                continue

            rate = float(row[header_map['rate']]) if row[header_map['rate']] is not None else 0.0
            if rate < 0:
                skipped_rows.append({
                    'row': row_idx,
                    'data': {
                        'item': item,
                        'qty': qty,
                        'rate': rate
                    },
                    'reason': 'Selling price cannot be negative'
                })
                continue

            discount = float(row[header_map.get('discount')]) if header_map.get('discount') is not None and row[header_map['discount']] is not None else 0.0

            sale_date = None
            if 'date' in header_map and row[header_map['date']] is not None:
                try:
                    if isinstance(row[header_map['date']], (int, float)):
                        from datetime import datetime, timedelta
                        sale_date = datetime(1899, 12, 30) + timedelta(days=row[header_map['date']])
                    else:
                        date_str = str(row[header_map['date']]).strip()
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                            try:
                                sale_date = datetime.strptime(date_str, fmt)
                                break
                            except:
                                continue
                        if sale_date is None:
                            sale_date = datetime.now()
                except:
                    sale_date = datetime.now()
            else:
                sale_date = datetime.now()

            rows_to_process.append({
                'row_idx': row_idx,
                'item': item,
                'qty': int(qty),
                'rate': rate,
                'discount': discount,
                'sale_date': sale_date
            })
        except Exception as e:
            skipped_rows.append({
                'row': row_idx,
                'data': {
                    'item': row[header_map.get('item')] if header_map.get('item') is not None else None,
                    'qty': row[header_map.get('qty')] if header_map.get('qty') is not None else None,
                    'rate': row[header_map.get('rate')] if header_map.get('rate') is not None else None
                },
                'reason': str(e)
            })

    total_rows = len(rows_to_process)
    update_job_progress(job_id, 
                        total=total_rows, 
                        processed=0, 
                        status='processing',
                        errors=error_rows,
                        result={'imported': 0, 'skipped': skipped_rows, 'message': 'Parsing completed'})

    if total_rows == 0:
        update_job_progress(job_id, status='done',
                            result={'imported': 0, 'skipped': skipped_rows, 'message': 'No valid rows to import'})
        return

    BATCH_SIZE = 5000
    imported_count = 0
    overall_errors = error_rows.copy()

    if not DATABASE_URL:
        update_job_progress(job_id, status='error', errors=['DATABASE_URL not configured'])
        return

    for i in range(0, total_rows, BATCH_SIZE):
        batch = rows_to_process[i:i+BATCH_SIZE]
        conn = None
        try:
            conn = psycopg2.connect(DATABASE_URL, sslmode='require')
            conn.autocommit = False
            cursor = conn.cursor()

            for entry in batch:
                # Find product
                cursor.execute(
                    "SELECT id, category FROM products WHERE name = %s",
                    (entry['item'],)
                )
                product = cursor.fetchone()
                if not product:
                    overall_errors.append(f"Row {entry['row_idx']}: Product '{entry['item']}' not found")
                    continue
                product_id, product_category = product
                if product_category != target_category:
                    overall_errors.append(f"Row {entry['row_idx']}: Product '{entry['item']}' category '{product_category}' != '{target_category}'")
                    continue

                subtotal = entry['qty'] * entry['rate']
                total = subtotal - entry['discount']
                cursor.execute("""
                    INSERT INTO sales (date, subtotal, discount, total, profit, reversed, payment_method)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (entry['sale_date'], subtotal, entry['discount'], total, 0, 0, 'cash'))
                sale_id = cursor.fetchone()[0]

                cursor.execute("""
                    SELECT id, cost_price, selling_price
                    FROM purchase_batches
                    WHERE product_id = %s AND remaining_quantity >= %s
                    ORDER BY date ASC
                    LIMIT 1
                """, (product_id, entry['qty']))
                batch_info = cursor.fetchone()
                if not batch_info:
                    overall_errors.append(f"Row {entry['row_idx']}: Insufficient stock for '{entry['item']}' (need {entry['qty']})")
                    cursor.execute("DELETE FROM sales WHERE id = %s", (sale_id,))
                    continue

                batch_id, cost_price, selling_price = batch_info
                selling_price = entry['rate']
                item_profit = (selling_price - cost_price) * entry['qty'] - entry['discount']

                cursor.execute("""
                    INSERT INTO sales_items
                    (sale_id, product_id, batch_id, quantity, selling_price, cost_price, profit, subtotal)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (sale_id, product_id, batch_id, entry['qty'], selling_price, cost_price, item_profit, subtotal))

                cursor.execute("""
                    UPDATE purchase_batches
                    SET remaining_quantity = remaining_quantity - %s
                    WHERE id = %s
                """, (entry['qty'], batch_id))

                cursor.execute("""
                    UPDATE products
                    SET stock = stock - %s
                    WHERE id = %s
                """, (entry['qty'], product_id))

                imported_count += 1
                if imported_count % 50 == 0:
                    update_job_progress(job_id, processed=imported_count)

            conn.commit()
            update_job_progress(job_id, processed=imported_count)

        except Exception as e:
            if conn:
                conn.rollback()
            overall_errors.append(f"Batch starting at row {batch[0]['row_idx']}: {str(e)}")
            update_job_progress(job_id, errors=overall_errors)
            break
        finally:
            if conn:
                conn.close()

    if overall_errors and len(overall_errors) > 0:
        update_job_progress(job_id, status='error', errors=overall_errors)
    else:
        update_job_progress(job_id, status='done',
                            result={
                                'imported': imported_count,
                                'skipped': skipped_rows,
                                'message': f'Imported {imported_count} sales records, {len(skipped_rows)} rows skipped'
                            })

@app.route('/api/sales/import', methods=['POST'])
@admin_required
def api_import_sales():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'File type not allowed'}), 400

    target_category = request.form.get('target_category', 'Accessory')
    if target_category not in ['Accessory', 'Screen']:
        return jsonify({'success': False, 'error': 'Invalid target category'}), 400

    file_content = file.read()
    file_stream = io.BytesIO(file_content)

    job_id = str(uuid.uuid4())

    # Insert initial job record
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO import_jobs (job_id, status, total, processed, errors, result)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (job_id, 'pending', 0, 0, '[]', None))
    conn.commit()
    conn.close()

    thread = threading.Thread(
        target=run_sales_import,
        args=(job_id, file_stream, target_category)
    )
    thread.daemon = True
    thread.start()

    return jsonify({'success': True, 'job_id': job_id})

# ===================== PROGRESS POLLING ENDPOINT =====================
@app.route('/api/import/progress/<job_id>', methods=['GET'])
@login_required
def api_import_progress(job_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT status, total, processed, errors, result, updated_at
        FROM import_jobs
        WHERE job_id = %s
    """, (job_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': 'Job not found'}), 404

    # Parse result if it's a JSON string (postgres returns JSONB as dict or list)
    result_data = row[4]
    if result_data and isinstance(result_data, str):
        try:
            result_data = json.loads(result_data)
        except:
            pass

    return jsonify({
        'success': True,
        'status': row[0],
        'total': row[1],
        'processed': row[2],
        'errors': row[3],
        'result': result_data,
        'updated_at': row[5].isoformat() if row[5] else None
    })
@app.route('/api/import/failed-report/<job_id>', methods=['GET'])
@login_required
def api_import_failed_report(job_id):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT result FROM import_jobs WHERE job_id = %s", (job_id,))
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': 'Job not found'}), 404

    result = row[0]
    # Ensure it's a dict
    if isinstance(result, str):
        try:
            result = json.loads(result)
        except:
            pass

    if not result or not result.get('skipped'):
        return jsonify({'success': False, 'error': 'No skipped rows to report'}), 400

    skipped = result['skipped']
    # ... rest of PDF generation is fine ...
    # Generate PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("📋 Import Failed Items Report", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Table header
    table_data = [["Row", "Name/Item", "Qty", "Rate", "Reason"]]
    for s in skipped:
        table_data.append([
            str(s.get('row', '')),
            s.get('data', {}).get('name') or s.get('data', {}).get('item') or '-',
            str(s.get('data', {}).get('qty', '')),
            str(s.get('data', {}).get('rate', '')),
            s.get('reason', '')
        ])

    table = Table(table_data, repeatRows=1, colWidths=[50, 150, 50, 50, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True,
                     download_name=f"failed_import_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                     mimetype='application/pdf')
@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({"status": "ok", "message": "Deployed version is current"})               
# ---------------------- RUN THE APP ----------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)